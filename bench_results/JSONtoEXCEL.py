#!/usr/bin/env python3
import argparse
from datetime import datetime
import json
import logging
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Set, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
try:
    import matplotlib
    matplotlib.use("Agg")
    from matplotlib import ticker as mticker
    from matplotlib import pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch
except ModuleNotFoundError:
    matplotlib = None
    mticker = plt = PdfPages = Line2D = Patch = None

_MPL_AVAILABLE = matplotlib is not None

TEST_CATEGORIES = [
    "AI",
    "Compression",
    "Cryptography_and_TLS",
    "Database",
    "FPU",
    "Java_Applications",
    "Memory_Access",
    "Multimedia",
    "Network",
    "Processor",
    "System",
    "Telecom",
]

# A-L columns as per spec
HEADERS = [
    "benchmark",          # A
    "test_snippet",       # B
    "test_name",          # C
    "os",                 # D
    "gcc_ver",            # E
    "thread",             # F
    "unit",               # G
    "machinename",        # H
    "cpu_name",           # I
    "score",              # J
    "relative_performance",  # K
    "performance",        # L
]

# Column widths (A-L)
_COL_WIDTHS = [50, 50, 30, 18, 18, 10, 15, 30, 35, 12, 20, 12]

# ---------------------------------------------------------------------------
# CSP Marker Table (Line Chart only)
# Determined from machinename prefix.  Sync with JSONtoEXCEL.md.
#
# | CSP   | prefix   | marker | fillstyle | shape |
# |-------|----------|--------|-----------|-------|
# | AWS   | aws-     | "o"    | "none"    | ○ 白抜き円 |
# | GCP   | gcp-     | "o"    | "full"    | ● 塗り円   |
# | OCI   | oci-     | "^"    | "none"    | △ 白抜き三角 |
# | Azure | azure-   | "^"    | "full"    | ▲ 塗り三角  |
# | Other | (else)   | "s"    | "full"    | ■ 正方形    |
# ---------------------------------------------------------------------------
CSP_MARKER_TABLE: Dict[str, Tuple[str, str]] = {
    "aws":   ("o", "none"),
    "gcp":   ("o", "full"),
    "oci":   ("^", "none"),
    "azure": ("^", "full"),
    "other": ("s", "full"),
}

_CSP_PREFIXES = [
    ("aws-",   "aws"),
    ("gcp-",   "gcp"),
    ("oci-",   "oci"),
    ("azure-", "azure"),
]

def detect_csp(machinename: str) -> str:
    """Return CSP key for machinename based on prefix matching."""
    name = machinename.casefold()
    for prefix, csp in _CSP_PREFIXES:
        if name.startswith(prefix):
            return csp
    return "other"


def find_comparison_block(payload: Dict[str, Any]) -> Dict[str, Any]:
    for key, value in payload.items():
        if key.endswith("_comparison") and isinstance(value, dict) and "workload" in value:
            return value
    raise ValueError("No *_comparison block with workload found")


def to_int_if_possible(value: Any) -> Any:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return value


def to_float_if_possible(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def is_rank_one(value: Any) -> bool:
    if value == 1:
        return True
    if isinstance(value, str) and value.strip() == "1":
        return True
    return False


def is_fallback_value(value: Any) -> bool:
    """Return True when value should be treated as missing/placeholder."""
    if value is None:
        return True
    if not isinstance(value, str):
        return False
    normalized = value.strip().lower()
    return normalized in {"", "n/a", "na", "none", "null"}


def extract_rows(payload: Dict[str, Any]) -> List[Tuple[Any, ...]]:
    """Extract 12-tuples:
    (benchmark, test_snippet, test_name, os, gcc_ver, thread, unit, machinename, cpu_name, score, relative, rank).
    """
    comparison = find_comparison_block(payload)
    workload = comparison.get("workload", {})

    rows: List[Tuple[Any, ...]] = []
    for benchmark_name, benchmark_group in workload.items():
        if not isinstance(benchmark_group, dict):
            continue

        # Current format:
        # workload[<benchmark>][test_snippet?, gcc_ver?, <test_name>][os][thread]...
        # Legacy format:
        # workload[<group>][<benchmark>][test_snippet?, gcc_ver?, <test_name>][os][thread]...
        # direct_nodes: (benchmark_name, test_snippet, gcc_ver, test_name, test_data)
        if "os" in benchmark_group or "test_snippet" in benchmark_group or "gcc_ver" in benchmark_group:
            direct_nodes = []
            direct_os_map = benchmark_group.get("os")
            fallback_snippet = benchmark_group.get("test_snippet", "N/A")
            fallback_gcc_ver = benchmark_group.get("gcc_ver", "14.2-system")
            if isinstance(direct_os_map, dict):
                direct_nodes.append((benchmark_name, fallback_snippet, fallback_gcc_ver, "", benchmark_group))
            else:
                for test_name, test_data in benchmark_group.items():
                    if not isinstance(test_data, dict):
                        continue
                    os_map = test_data.get("os")
                    if not isinstance(os_map, dict):
                        continue
                    test_snippet = test_data.get("test_snippet", fallback_snippet)
                    gcc_ver = test_data.get("gcc_ver", fallback_gcc_ver)
                    if is_fallback_value(test_snippet):
                        test_snippet = fallback_snippet
                    if is_fallback_value(gcc_ver):
                        gcc_ver = fallback_gcc_ver
                    direct_nodes.append((benchmark_name, test_snippet, gcc_ver, test_name, test_data))
        else:
            direct_nodes = []
            for nested_benchmark_name, benchmark_data in benchmark_group.items():
                if not isinstance(benchmark_data, dict):
                    continue
                fallback_snippet = benchmark_data.get("test_snippet", "N/A")
                fallback_gcc_ver = benchmark_data.get("gcc_ver", "14.2-system")
                direct_os_map = benchmark_data.get("os")
                if isinstance(direct_os_map, dict):
                    direct_nodes.append((nested_benchmark_name, fallback_snippet, fallback_gcc_ver, "", benchmark_data))
                    continue
                for test_name, test_data in benchmark_data.items():
                    if not isinstance(test_data, dict):
                        continue
                    os_map = test_data.get("os")
                    if not isinstance(os_map, dict):
                        continue
                    test_snippet = test_data.get("test_snippet", fallback_snippet)
                    gcc_ver = test_data.get("gcc_ver", fallback_gcc_ver)
                    if is_fallback_value(test_snippet):
                        test_snippet = fallback_snippet
                    if is_fallback_value(gcc_ver):
                        gcc_ver = fallback_gcc_ver
                    direct_nodes.append((nested_benchmark_name, test_snippet, gcc_ver, test_name, test_data))

        for active_benchmark_name, test_snippet, gcc_ver, test_name, test_node in direct_nodes:
            os_map = test_node.get("os", {})
            for os_name, os_data in os_map.items():
                thread_map = os_data.get("thread", {}) if isinstance(os_data, dict) else {}
                for thread, thread_data in thread_map.items():
                    if not isinstance(thread_data, dict):
                        continue

                    unit = thread_data.get("unit")
                    ranking = thread_data.get("leaderboard") or thread_data.get("ranking") or []

                    for item in ranking:
                        if not isinstance(item, dict):
                            continue

                        score = to_float_if_possible(item.get("score", item.get("efficiency_score")))
                        relative = item.get(
                            "relative_performance",
                            item.get("relative_cost_efficiency"),
                        )
                        relative = to_float_if_possible(relative)
                        rank = item.get("rank")

                        rows.append(
                            (
                                active_benchmark_name,
                                test_snippet,
                                test_name,
                                os_name,
                                gcc_ver,
                                to_int_if_possible(thread),
                                unit,
                                item.get("machinename"),
                                item.get("cpu_name"),
                                score,
                                relative,
                                rank,  # index 11: rank (replaced by performance later)
                            )
                        )

    return rows


def add_original_column(rows: Iterable[Tuple[Any, ...]], json_path: Path) -> List[Tuple[Any, ...]]:
    """Replace rank (index 11) with L-column performance value.

    Result tuple: (benchmark, test_snippet, test_name, os, gcc_ver, thread, unit, machinename, cpu_name, score, relative, performance)
    """
    row_list = list(rows)
    is_performance_analysis = "performance_analysis" in json_path.name

    if not is_performance_analysis:
        return [
            (
                benchmark,
                test_snippet,
                test_name,
                os_name,
                gcc_ver,
                thread,
                unit,
                machine_name,
                cpu_name,
                score,
                relative,
                None,
            )
            for (
                benchmark,
                test_snippet,
                test_name,
                os_name,
                gcc_ver,
                thread,
                unit,
                machine_name,
                cpu_name,
                score,
                relative,
                _,
            ) in row_list
        ]

    # Find baseline: rank=1, minimum thread, per (benchmark, test_name, os, unit)
    baseline_scores: Dict[Tuple[Any, Any, Any, Any], float] = {}
    selected_threads: Dict[Tuple[Any, Any, Any, Any], int] = {}

    for row in row_list:
        (
            benchmark,
            test_snippet,
            test_name,
            os_name,
            gcc_ver,
            thread,
            unit,
            _,
            _,
            score,
            _,
            rank,
        ) = row
        if not isinstance(thread, int):
            continue
        if not isinstance(score, (int, float)):
            continue
        if not is_rank_one(rank):
            continue

        key = (benchmark, test_name, os_name, unit)
        prev_thread = selected_threads.get(key)
        if prev_thread is None or thread < prev_thread:
            selected_threads[key] = thread
            baseline_scores[key] = float(score)

    enriched_rows: List[Tuple[Any, ...]] = []
    for row in row_list:
        (
            benchmark,
            test_snippet,
            test_name,
            os_name,
            gcc_ver,
            thread,
            unit,
            machine_name,
            cpu_name,
            score,
            relative,
            _,
        ) = row
        key = (benchmark, test_name, os_name, unit)
        baseline = baseline_scores.get(key)

        performance = None
        if baseline not in (None, 0) and isinstance(score, (int, float)):
            unit_normalized = unit.strip().lower() if isinstance(unit, str) else ""
            if unit_normalized == "microseconds":
                performance = (baseline / float(score)) * 100 if score != 0 else None
            else:
                performance = (float(score) / baseline) * 100

        enriched_rows.append(
            (
                benchmark,
                test_snippet,
                test_name,
                os_name,
                gcc_ver,
                thread,
                unit,
                machine_name,
                cpu_name,
                score,
                relative,
                performance,
            )
        )

    return enriched_rows


# ---------------------------------------------------------------------------
# Sorting helpers (shared by build_output_rows and graph)
# ---------------------------------------------------------------------------

def _normalize_sort_text(value: Any, case_insensitive: bool = True, collapse_spaces: bool = True) -> str:
    if value is None:
        return ""
    text = str(value)
    if collapse_spaces:
        text = " ".join(text.strip().split())
    return text.casefold() if case_insensitive else text


def _natural_sort_key(
    value: Any,
    case_insensitive: bool = True,
    collapse_spaces: bool = True,
) -> Tuple[Tuple[int, Any], ...]:
    text = _normalize_sort_text(value, case_insensitive=case_insensitive, collapse_spaces=collapse_spaces)
    parts = re.split(r"(\d+)", text)
    key_parts: List[Tuple[int, Any]] = []
    for part in parts:
        if part == "":
            continue
        if part.isdigit():
            key_parts.append((0, int(part)))
        else:
            key_parts.append((1, part))
    return tuple(key_parts)


def _thread_sort_value(value: Any) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return 10**9


def build_output_rows(rows: Iterable[Tuple[Any, ...]]) -> List[Tuple[Any, ...]]:
    """Sort 12-tuples by benchmark (A) → machinename (H) → thread (F), natural order."""
    row_list = list(rows)

    def sort_key(item: Tuple[Any, ...]) -> Tuple:
        benchmark = item[0]   # A
        machine_name = item[7]  # H
        thread = item[5]  # F
        return (
            _natural_sort_key(benchmark, case_insensitive=True, collapse_spaces=False),
            _natural_sort_key(machine_name, case_insensitive=True, collapse_spaces=True),
            _thread_sort_value(thread),
        )

    row_list.sort(key=sort_key)
    return row_list


def write_xlsx(rows: Iterable[Tuple[Any, ...]], output_path: Path) -> None:
    """Write 12-column Excel (A-L) per spec."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    # Header row (bold)
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    # Column widths
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[col_letter].width = width

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            # J=10 (score), K=11 (relative_performance), L=12 (performance) → 2 decimal places
            if col_idx in (10, 11, 12) and isinstance(value, (int, float)):
                cell.number_format = "0.00"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(output_path))


def read_rows_from_xlsx(xlsx_path: Path) -> List[Tuple[Any, ...]]:
    """Read 12-column Excel back as tuples for --graph mode."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows: List[Tuple[Any, ...]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        values = list(row)
        if len(values) < 12:
            values.extend([None] * (12 - len(values)))
        rows.append(tuple(values[:12]))
    wb.close()
    return rows


def generate_graph_pdf(rows: Iterable[Tuple[Any, ...]], output_path: Path) -> int:
    """Generate per-(benchmark, test_name) PDF graphs.  Expects 12-tuples:
    (benchmark, test_snippet, test_name, os, gcc_ver, thread, unit, machinename, cpu_name, score, relative, performance).

    One page per (benchmark, test_name) combination.
    Graph title layout:
      Main title  — test_snippet from row data (e.g. "SVT-AV1-4.0")
      Subtitle    — test_name (unit)[, Thread=N]
    """
    if not _MPL_AVAILABLE:
        logging.warning("matplotlib is not available; skipping PDF generation: %s", output_path)
        return 0

    # Key: (benchmark, test_name)
    grouped: Dict[Tuple[str, str], Dict[str, List[Tuple[int, float]]]] = {}
    snippet_map: Dict[Tuple[str, str], str] = {}
    benchmark_units: Dict[Tuple[str, str], str] = {}
    machine_family_by_key: Dict[Tuple[str, str], Dict[str, str]] = {}

    def detect_machine_family(cpu_name: Any) -> str:
        cpu_text = cpu_name.casefold() if isinstance(cpu_name, str) else ""
        arm_markers = ("arm", "aarch64", "graviton", "ampere", "neoverse", "cortex")
        amd_markers = ("amd", "epyc")
        intel_markers = ("intel", "xeon")
        if any(m in cpu_text for m in arm_markers):
            return "arm"
        if any(m in cpu_text for m in amd_markers):
            return "amd"
        if any(m in cpu_text for m in intel_markers):
            return "intel"
        return "other"

    for row in rows:
        (
            benchmark,
            test_snippet,
            test_name,
            _,
            _,
            thread,
            unit,
            machinename,
            cpu_name,
            _,
            _,
            performance,
        ) = row
        if not isinstance(benchmark, str):
            continue
        if not isinstance(machinename, str):
            continue
        if not isinstance(thread, int):
            continue
        if not isinstance(performance, (int, float)):
            continue

        key = (benchmark, test_name if isinstance(test_name, str) else "")
        if key not in benchmark_units and isinstance(unit, str):
            benchmark_units[key] = unit

        benchmark_map = grouped.setdefault(key, {})
        benchmark_map.setdefault(machinename, []).append((thread, float(performance)))
        family_map = machine_family_by_key.setdefault(key, {})
        family_map.setdefault(machinename, detect_machine_family(cpu_name))
        if isinstance(test_snippet, str) and not is_fallback_value(test_snippet):
            snippet_map[key] = test_snippet

    output_path.parent.mkdir(parents=True, exist_ok=True)
    page_count = 0

    def shorten_machinename(name: str, max_len: int = 18) -> str:
        text = name.strip()
        if len(text) <= max_len:
            return text
        return text[: max_len - 1] + "\u2026"

    def build_machine_colors(key: Tuple[str, str], machine_names: List[str]) -> Dict[str, str]:
        palettes = {
            "arm":   ["#1f77b4", "#4f97cc", "#76afd8", "#9bc8e7", "#c2dff2"],
            "intel": ["#b22222", "#c74444", "#d96868", "#e58d8d", "#f0b3b3"],
            "amd":   ["#111111", "#2a2a2a", "#444444", "#5e5e5e", "#7a7a7a"],
            "other": ["#666666", "#808080", "#9a9a9a", "#b3b3b3", "#cdcdcd"],
        }
        family_counters: Dict[str, int] = {"arm": 0, "intel": 0, "amd": 0, "other": 0}
        family_map = machine_family_by_key.get(key, {})
        machine_colors: Dict[str, str] = {}
        for machinename in machine_names:
            family = family_map.get(machinename, "other")
            if family not in palettes:
                family = "other"
            palette = palettes[family]
            index = family_counters[family] % len(palette)
            machine_colors[machinename] = palette[index]
            family_counters[family] += 1
        return machine_colors

    with PdfPages(str(output_path)) as pdf:
        if not grouped:
            figure, axis = plt.subplots(figsize=(11.69, 8.27))
            axis.axis("off")
            axis.text(
                0.5, 0.5,
                "No performance data available for graph generation",
                ha="center", va="center", fontsize=14,
            )
            pdf.savefig(figure)
            plt.close(figure)
            return 1

        for key, machine_map in grouped.items():
            benchmark, test_name = key
            figure, axis = plt.subplots(figsize=(11.69, 8.27))
            figure.subplots_adjust(right=0.72)
            legend_handles: List[Any] = []
            all_threads = sorted(
                {thread for points in machine_map.values() for thread, _ in points}
            )
            single_thread_mode = len(all_threads) == 1

            machine_items = sorted(machine_map.items(), key=lambda item: _natural_sort_key(item[0]))
            machine_names = [name for name, _ in machine_items]
            machine_colors = build_machine_colors(key, machine_names)

            if single_thread_mode:
                # Bar chart: sorted by performance descending (highest = leftmost)
                ranked_items: List[Tuple[str, List[Tuple[int, float]], float]] = []
                for machinename, points in machine_items:
                    y_values = [v for _, v in points]
                    if not y_values:
                        continue
                    bar_value = sum(y_values) / len(y_values)
                    ranked_items.append((machinename, points, bar_value))
                ranked_items.sort(key=lambda item: item[2], reverse=True)

                x_positions = list(range(len(ranked_items)))
                x_labels: List[str] = []
                label_counts: Dict[str, int] = {}

                for index, (machinename, _, bar_value) in enumerate(ranked_items):
                    short_label = shorten_machinename(machinename)
                    label_counts[short_label] = label_counts.get(short_label, 0) + 1
                    if label_counts[short_label] > 1:
                        short_label = f"{short_label}#{label_counts[short_label]}"
                    x_labels.append(short_label)
                    color = machine_colors.get(machinename, "#808080")
                    axis.bar(index, bar_value, width=0.7, color=color)
                    legend_handles.append(Patch(facecolor=color, edgecolor=color, label=machinename))

                axis.set_xticks(x_positions)
                axis.set_xticklabels(x_labels)
                axis.set_xlabel("machinename")
                axis.tick_params(axis="x", labelrotation=25)
                max_bar_value = max((item[2] for item in ranked_items), default=0.0)
                axis.set_ylim(bottom=0, top=max(100.0, max_bar_value * 1.05 or 100.0))
                axis.yaxis.set_major_locator(mticker.MaxNLocator(nbins=8, prune="both"))

            else:
                # Line chart with series numbers alternating left/right
                line_styles = ["-", "--", ":", "-."]
                for series_number, (machinename, points) in enumerate(machine_items, start=1):
                    points_sorted = sorted(points, key=lambda item: item[0])
                    x_values = [item[0] for item in points_sorted]
                    y_values = [item[1] for item in points_sorted]
                    legend_label = f"{series_number}: {machinename}"
                    color = machine_colors.get(machinename, "#808080")
                    line_style = line_styles[(series_number - 1) % len(line_styles)]
                    csp_key = detect_csp(machinename)
                    csp_marker, csp_fillstyle = CSP_MARKER_TABLE.get(csp_key, CSP_MARKER_TABLE["other"])
                    axis.plot(
                        x_values, y_values,
                        marker=csp_marker, fillstyle=csp_fillstyle, linewidth=1.5,
                        label=legend_label, color=color, linestyle=line_style,
                    )
                    legend_handles.append(
                        Line2D([0], [0], color=color, linestyle=line_style,
                               marker=csp_marker, fillstyle=csp_fillstyle,
                               linewidth=1.5, label=legend_label)
                    )

                    if x_values and y_values and len(machine_items) <= 20:
                        place_on_left = series_number % 2 == 1
                        anchor_index = 0 if place_on_left else -1
                        x_anchor = float(x_values[anchor_index])
                        y_anchor = float(y_values[anchor_index])
                        x_offset = -7 if place_on_left else 7
                        align = "right" if place_on_left else "left"
                        y_offset = ((series_number - 1) % 4 - 1.5) * 4
                        axis.annotate(
                            str(series_number),
                            xy=(x_anchor, y_anchor),
                            xytext=(x_offset, y_offset),
                            textcoords="offset points",
                            ha=align, va="center",
                            fontsize=8, fontweight="bold", color=color,
                            bbox={"boxstyle": "round,pad=0.1", "fc": "white", "ec": "none", "alpha": 0.7},
                        )

                axis.set_xlabel("thread")
                # Keep chart readable even when performance values span very wide ranges.
                # A dense fixed step (e.g. 10) causes 1000+ ticks and unreadable labels.
                axis.yaxis.set_major_locator(mticker.MaxNLocator(nbins=8, prune="both"))

            # Common axes/title
            unit_label = benchmark_units.get(key)
            if single_thread_mode and all_threads:
                subtitle = (
                    f"{test_name} ({unit_label}), Thread={all_threads[0]}"
                    if unit_label else f"{test_name}, Thread={all_threads[0]}"
                )
            else:
                subtitle = f"{test_name} ({unit_label})" if unit_label else test_name
            main_title = snippet_map.get(key, "")
            if not is_fallback_value(main_title):
                figure.suptitle(main_title, x=0.5, y=0.98, ha="center", va="top", fontsize=13, fontweight="bold")
            axis.set_title(subtitle, fontsize=9, pad=2)
            axis.set_ylabel("performance")
            axis.grid(True, axis="y", linestyle="--", alpha=0.35)

            if legend_handles:
                axis.legend(
                    handles=legend_handles,
                    loc="upper left",
                    bbox_to_anchor=(1.01, 1.0),
                    borderaxespad=0.0,
                    frameon=False,
                    fontsize=7,
                )

            pdf.savefig(figure)
            plt.close(figure)
            page_count += 1

    return page_count


def convert_json_file(json_path: Path, output_ext: str) -> int:
    with json_path.open("r", encoding="utf-8") as fp:
        payload = json.load(fp)

    raw_rows = extract_rows(payload)
    rows_with_performance = add_original_column(raw_rows, json_path)
    output_rows = build_output_rows(rows_with_performance)
    output_path = json_path.with_suffix(output_ext)
    pdf_path = json_path.with_suffix(".pdf")

    write_xlsx(output_rows, output_path)

    is_performance_analysis = "performance_analysis" in json_path.name
    if is_performance_analysis:
        generate_graph_pdf(output_rows, pdf_path)
        if not _MPL_AVAILABLE and pdf_path.exists():
            logging.warning("PDF generation skipped because matplotlib is missing for %s", json_path.name)
    elif pdf_path.exists():
        pdf_path.unlink()

    return len(output_rows)


def gather_target_jsons(global_dir: Path) -> List[Path]:
    json_files: List[Path] = []
    for category in TEST_CATEGORIES:
        category_dir = global_dir / category
        if not category_dir.exists() or not category_dir.is_dir():
            continue
        json_files.extend(
            sorted(
                json_file
                for json_file in category_dir.glob("*.json")
                if json_file.name.startswith(f"{category}_")
            )
        )
    # Fallback: if no per-category subdirs found, look for flat *performance_analysis*.json
    # files directly in global_dir (e.g. global_performance_analysis.json)
    if not json_files:
        json_files.extend(
            sorted(
                json_file
                for json_file in global_dir.glob("*.json")
                if "performance_analysis" in json_file.name
            )
        )
    return json_files


def graph_mode(global_dir: Path) -> int:
    """Read existing performance_analysis Excel files and regenerate PDFs."""
    if not _MPL_AVAILABLE:
        logging.warning("matplotlib is not available; graph mode skipped")
        return 0

    errors = 0
    generated = 0
    for category in TEST_CATEGORIES:
        category_dir = global_dir / category
        if not category_dir.exists() or not category_dir.is_dir():
            continue
        for xlsx_path in sorted(category_dir.glob("*.xlsx")):
            if "performance_analysis" not in xlsx_path.name:
                continue
            pdf_path = xlsx_path.with_suffix(".pdf")
            try:
                rows = read_rows_from_xlsx(xlsx_path)
                if not rows:
                    logging.warning("[WARN] No data in %s", xlsx_path)
                    continue
                count = generate_graph_pdf(rows, pdf_path)
                logging.info("[OK] %s -> %s (%d pages)", xlsx_path, pdf_path, count)
                generated += 1
            except Exception as exc:
                logging.error("[NG] %s: %s", xlsx_path, exc)
                errors += 1
    logging.info("Graph mode: %d PDF(s) generated, %d error(s)", generated, errors)
    return errors


def _sort_strings_natural(values: Iterable[str]) -> List[str]:
    return sorted(values, key=lambda value: _natural_sort_key(value, case_insensitive=True, collapse_spaces=True))


def build_coverage_nog_for_category(json_path: Path) -> Dict[str, Any]:
    with json_path.open("r", encoding="utf-8") as fp:
        payload = json.load(fp)

    rows = extract_rows(payload)
    testcategory = json_path.stem.replace("_performance_analysis", "")
    machine_population: Set[str] = set()
    presence: Dict[Tuple[str, int, str], Set[str]] = {}
    benchmark_threads: Dict[Tuple[str, int], Set[str]] = {}

    for row in rows:
        benchmark = row[0]
        test_name = row[2]
        thread = row[5]
        machinename = row[7]
        if not isinstance(benchmark, str):
            continue
        if not isinstance(thread, int):
            continue
        if not isinstance(machinename, str):
            continue
        if machinename.strip() == "":
            continue
        if not isinstance(test_name, str):
            test_name = ""
        machine_population.add(machinename)
        key = (benchmark, thread, test_name)
        presence.setdefault(key, set()).add(machinename)
        benchmark_threads.setdefault((benchmark, thread), set()).add(test_name)

    nog_by_benchmark: Dict[str, List[Dict[str, Any]]] = {}
    nog_count = 0
    all_machines = set(machine_population)

    for benchmark, thread in sorted(
        benchmark_threads.keys(),
        key=lambda item: (
            _natural_sort_key(item[0], case_insensitive=True, collapse_spaces=False),
            item[1],
        ),
    ):
        missing_machines_union: Set[str] = set()
        missing_test_names: Set[str] = set()
        for test_name in benchmark_threads[(benchmark, thread)]:
            found = presence.get((benchmark, thread, test_name), set())
            missing_here = all_machines - found
            if missing_here:
                missing_machines_union.update(missing_here)
                if test_name.strip() != "":
                    missing_test_names.add(test_name)

        if not missing_machines_union:
            continue

        csp_breakdown: Dict[str, List[str]] = {}
        for machinename in _sort_strings_natural(missing_machines_union):
            csp = detect_csp(machinename)
            csp_breakdown.setdefault(csp, []).append(machinename)

        nog_by_benchmark.setdefault(benchmark, []).append(
            {
                "thread": str(thread),
                "missing_count": len(missing_machines_union),
                "missing_test_names": _sort_strings_natural(missing_test_names),
                "csp_breakdown": csp_breakdown,
            }
        )
        nog_count += 1

    for benchmark in list(nog_by_benchmark.keys()):
        nog_by_benchmark[benchmark].sort(key=lambda item: _thread_sort_value(item.get("thread")))

    return {
        testcategory: {
            "nog_count": nog_count,
            "nog": nog_by_benchmark,
        }
    }


def generate_coverage_nog_report(global_dir: Path) -> Dict[str, Any]:
    categories: Dict[str, Any] = {}
    for category in TEST_CATEGORIES:
        json_path = global_dir / category / f"{category}_performance_analysis.json"
        if not json_path.exists():
            continue
        if not json_path.is_file():
            continue
        categories.update(build_coverage_nog_for_category(json_path))

    return {
        "schema_version": "1.0",
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "results": {
            "testcategory": categories,
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert benchmark JSON files to Excel")
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Workspace root path (default: script directory)",
    )
    parser.add_argument(
        "--ext",
        choices=[".xlsx"],
        default=".xlsx",
        help="Output extension (default: .xlsx)",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=None,
        help="Log file directory (default: $PWD/log)",
    )
    parser.add_argument(
        "--graph",
        action="store_true",
        help="Read existing Excel files and regenerate PDFs for performance_analysis files",
    )
    parser.add_argument(
        "--coverage-out",
        type=Path,
        default=None,
        help='Coverage NOG JSON output path (default: "<root>/coverage_nog_all_<timestamp>.json")',
    )
    args = parser.parse_args()

    log_dir: Path = args.log if args.log is not None else Path.cwd() / "log"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "JSONtoEXCEL.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stderr),
        ],
    )

    global_dir = args.root / "global"

    if args.graph:
        errors = graph_mode(global_dir)
        return 0 if errors == 0 else 2

    targets = gather_target_jsons(global_dir)
    if not targets:
        logging.warning("No target JSON files found under %s", global_dir)
        return 1

    total_rows = 0
    converted = 0

    for json_file in targets:
        try:
            row_count = convert_json_file(json_file, args.ext)
            total_rows += row_count
            converted += 1
            logging.info("[OK] %s -> %s (%d rows)", json_file, json_file.with_suffix(args.ext), row_count)
        except Exception as exc:
            logging.error("[NG] %s: %s", json_file, exc)

    coverage_report = generate_coverage_nog_report(global_dir)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    coverage_path = args.coverage_out if args.coverage_out is not None else args.root / f"coverage_nog_all_{timestamp}.json"
    coverage_path.parent.mkdir(parents=True, exist_ok=True)
    with coverage_path.open("w", encoding="utf-8") as fp:
        json.dump(coverage_report, fp, ensure_ascii=False, indent=2)
        fp.write("\n")
    print(json.dumps(coverage_report, ensure_ascii=False, indent=2))
    logging.info("Coverage NOG JSON: %s", coverage_path)

    logging.info("Converted: %d/%d files, %d total rows", converted, len(targets), total_rows)
    return 0 if converted == len(targets) else 2


if __name__ == "__main__":
    raise SystemExit(main())
